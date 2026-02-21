#!/usr/bin/env python3
"""
PKD Literature Search Script
Searches PubMed, bioRxiv, and medRxiv for polycystic kidney disease papers
Generates Excel and PDF reports with comprehensive citations

Usage (CLI):
    python pkd_literature_search.py --start 2026-02-01 --end 2026-02-08
    python pkd_literature_search.py  # Uses last 7 days by default

Usage (Streamlit):
    streamlit run app.py
"""

import argparse
import io
import time
from datetime import datetime, timedelta
from typing import List, Dict, Tuple, Optional

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY


class PKDLiteratureSearch:
    """Comprehensive PKD literature search across multiple databases.

    Supports both file-path and in-memory (BytesIO) report generation so the
    same class can power a CLI script *and* a Streamlit web app.
    """

    def __init__(self, start_date: str, end_date: str, output_dir: str = "."):
        self.start_date = start_date
        self.end_date = end_date
        self.output_dir = output_dir
        self.all_papers: List[Dict] = []
        self.pubmed_base = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"

    # ------------------------------------------------------------------
    # PubMed
    # ------------------------------------------------------------------

    def search_pubmed(self) -> List[str]:
        """Search PubMed for PKD papers and return a list of PMIDs."""
        print(f"Searching PubMed for papers from {self.start_date} to {self.end_date}...")

        query = (
            '("polycystic kidney" OR "polycystic kidney disease" '
            'OR "ADPKD" OR "ARPKD" OR PKD1 OR PKD2)'
        )

        params = {
            "db": "pubmed",
            "term": query,
            "retmax": 1000,
            "retmode": "json",
            "datetype": "pdat",
            "mindate": self.start_date.replace("-", "/"),
            "maxdate": self.end_date.replace("-", "/"),
        }

        try:
            response = requests.get(
                f"{self.pubmed_base}esearch.fcgi", params=params, timeout=30
            )
            response.raise_for_status()
            data = response.json()
            pmids = data.get("esearchresult", {}).get("idlist", [])
            print(f"  Found {len(pmids)} papers in PubMed")
            return pmids
        except Exception as e:
            print(f"  Error searching PubMed: {e}")
            return []

    def get_pubmed_metadata(
        self, pmids: List[str], batch_size: int = 100
    ) -> List[Dict]:
        """Retrieve metadata for PubMed articles in batches with retry logic."""
        print(f"Retrieving metadata for {len(pmids)} PubMed articles...")
        all_metadata: List[Dict] = []

        for i in range(0, len(pmids), batch_size):
            batch = pmids[i : i + batch_size]
            batch_num = (i // batch_size) + 1
            total_batches = (len(pmids) + batch_size - 1) // batch_size
            print(f"  Processing batch {batch_num}/{total_batches} ({len(batch)} PMIDs)...")

            max_retries = 3
            for attempt in range(max_retries):
                try:
                    summary_params = {
                        "db": "pubmed",
                        "id": ",".join(batch),
                        "retmode": "json",
                    }
                    resp = requests.get(
                        f"{self.pubmed_base}esummary.fcgi",
                        params=summary_params,
                        timeout=30,
                    )
                    resp.raise_for_status()
                    summary_data = resp.json()

                    for pmid in batch:
                        if pmid in summary_data.get("result", {}):
                            article = summary_data["result"][pmid]
                            all_authors = [
                                a.get("name", "") for a in article.get("authors", [])
                            ]
                            authors_string = ", ".join(all_authors) if all_authors else ""

                            metadata = {
                                "pmid": pmid,
                                "title": article.get("title", ""),
                                "authors": authors_string,
                                "journal": article.get("source", ""),
                                "year": (
                                    article.get("pubdate", "").split()[0]
                                    if article.get("pubdate")
                                    else ""
                                ),
                                "doi": next(
                                    (
                                        aid["value"]
                                        for aid in article.get("articleids", [])
                                        if aid.get("idtype") == "doi"
                                    ),
                                    "",
                                ),
                                "source": "PubMed",
                            }
                            all_metadata.append(metadata)

                    print(f"    Batch {batch_num} completed successfully")
                    break
                except Exception as e:
                    if attempt < max_retries - 1:
                        wait_time = (attempt + 1) * 2
                        print(f"    Attempt {attempt + 1} failed: {e}")
                        print(f"    Retrying in {wait_time} seconds...")
                        time.sleep(wait_time)
                    else:
                        print(
                            f"    Batch {batch_num} failed after {max_retries} attempts: {e}"
                        )

            time.sleep(0.5)

        print(f"  Successfully retrieved metadata for {len(all_metadata)} articles")
        return all_metadata

    # ------------------------------------------------------------------
    # bioRxiv / medRxiv
    # ------------------------------------------------------------------

    PKD_KEYWORDS = [
        "polycystic kidney",
        "polycystic kidney disease",
        "adpkd",
        "arpkd",
        "pkd1",
        "pkd2",
        "pkhd1",
        "autosomal dominant polycystic",
        "autosomal recessive polycystic",
        "kidney cyst",
        "renal cyst",
        "cystogenesis",
        "polycystin",
        "fibrocystin",
        "polyductin",
    ]

    def _is_pkd_relevant(self, title: str, abstract: str = "") -> bool:
        text = (title + " " + abstract).lower()
        return any(kw in text for kw in self.PKD_KEYWORDS)

    def _search_preprint_server(self, server: str) -> List[Dict]:
        """Search bioRxiv or medRxiv for PKD preprints via the public API."""
        base_url = f"https://api.biorxiv.org/details/{server}"
        papers: List[Dict] = []
        cursor = 0
        page_size = 100
        seen_dois: set = set()

        while True:
            url = f"{base_url}/{self.start_date}/{self.end_date}/{cursor}"

            max_retries = 3
            data = None
            for attempt in range(max_retries):
                try:
                    resp = requests.get(url, timeout=30)
                    resp.raise_for_status()
                    data = resp.json()
                    break
                except Exception as e:
                    if attempt < max_retries - 1:
                        wait_time = (attempt + 1) * 2
                        print(f"    Attempt {attempt + 1} failed at cursor {cursor}: {e}")
                        print(f"    Retrying in {wait_time} seconds...")
                        time.sleep(wait_time)
                    else:
                        print(
                            f"    API error at cursor {cursor} after {max_retries} attempts: {e}"
                        )

            if data is None:
                break

            collection = data.get("collection", [])
            if not collection:
                break

            for item in collection:
                title = item.get("title", "")
                abstract = item.get("abstract", "")
                doi = item.get("doi", "")
                if self._is_pkd_relevant(title, abstract):
                    if doi and doi in seen_dois:
                        continue
                    if doi:
                        seen_dois.add(doi)

                    papers.append(
                        {
                            "pmid": "",
                            "title": title,
                            "authors": item.get("authors", ""),
                            "journal": f"{server} (preprint)",
                            "year": item.get("date", "")[:4],
                            "doi": doi,
                            "source": server,
                        }
                    )

            total_msg = data.get("messages", [{}])[0]
            total_count = int(total_msg.get("total", 0)) if total_msg else 0
            cursor += page_size
            if cursor >= total_count:
                break

            time.sleep(0.3)

        return papers

    def search_biorxiv(self) -> List[Dict]:
        print(f"Searching bioRxiv for preprints from {self.start_date} to {self.end_date}...")
        papers = self._search_preprint_server("biorxiv")
        print(f"  Found {len(papers)} PKD-relevant preprints in bioRxiv")
        return papers

    def search_medrxiv(self) -> List[Dict]:
        print(f"Searching medRxiv for preprints from {self.start_date} to {self.end_date}...")
        papers = self._search_preprint_server("medrxiv")
        print(f"  Found {len(papers)} PKD-relevant preprints in medRxiv")
        return papers

    # ------------------------------------------------------------------
    # Categorisation helpers
    # ------------------------------------------------------------------

    def categorize_papers(self, papers: List[Dict]) -> Dict[str, List[Dict]]:
        categories: Dict[str, List[Dict]] = {
            "genetics": [],
            "therapeutics": [],
            "metabolism": [],
            "pathophysiology": [],
            "clinical": [],
            "cross_species": [],
            "dataset": [],
            "other": [],
        }

        for paper in papers:
            t = paper.get("title", "").lower()
            if any(w in t for w in ["genetic", "mutation", "variant", "sequencing", "gene"]):
                categories["genetics"].append(paper)
            elif any(w in t for w in ["drug", "therapeutic", "treatment", "inhibitor", "trial"]):
                categories["therapeutics"].append(paper)
            elif any(w in t for w in ["metabol", "mitochondr", "cholesterol", "amino acid"]):
                categories["metabolism"].append(paper)
            elif any(w in t for w in ["mouse", "mice", "rat", "model", "crispr"]):
                categories["cross_species"].append(paper)
            elif any(w in t for w in ["cohort", "registry", "dataset", "population"]):
                categories["dataset"].append(paper)
            elif any(w in t for w in ["pathophysiology", "mechanism", "pathway"]):
                categories["pathophysiology"].append(paper)
            elif any(w in t for w in ["patient", "clinical", "case"]):
                categories["clinical"].append(paper)
            else:
                categories["other"].append(paper)

        return categories

    # ------------------------------------------------------------------
    # Text helpers
    # ------------------------------------------------------------------

    @staticmethod
    def create_summary(title: str, max_words: int = 8) -> str:
        words = title.split()
        summary_words = words[:max_words] if len(words) <= max_words else words[: max_words - 1]
        summary = " ".join(summary_words)
        if len(words) > max_words:
            summary = summary.rstrip(".,;:") + "..."
        return summary

    @staticmethod
    def create_key_findings(title: str, max_words: int = 20) -> str:
        words = title.replace(":", " ").split()
        skip = {"the", "a", "an", "in", "on", "at", "to", "for", "of", "and", "with"}
        filtered = [
            w
            for w in words
            if w.lower() not in skip
            or len([x for x in words if x.lower() not in skip]) < 10
        ]
        return " ".join(filtered[:max_words])

    @staticmethod
    def extract_last_author_name(authors_string: str) -> str:
        if not authors_string:
            return ""
        if ";" in authors_string:
            authors = [a.strip() for a in authors_string.split(";")]
        else:
            authors = [a.strip() for a in authors_string.split(",")]
        authors = [a for a in authors if a]
        if not authors:
            return ""
        parts = authors[-1].split()
        return parts[0] if parts else ""

    # ------------------------------------------------------------------
    # Link helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _paper_link(paper: Dict) -> str:
        doi = paper.get("doi", "")
        pmid = paper.get("pmid", "")
        if doi:
            return f"https://doi.org/{doi}"
        if pmid:
            return f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
        return ""

    @staticmethod
    def _paper_link_label(paper: Dict) -> str:
        doi = paper.get("doi", "")
        pmid = paper.get("pmid", "")
        link = PKDLiteratureSearch._paper_link(paper)
        if not link:
            return ""
        if pmid:
            return f'<link href="{link}" color="blue">PMID: {pmid}</link>'
        if doi:
            return f'<link href="{link}" color="blue">DOI: {doi}</link>'
        return ""

    # ------------------------------------------------------------------
    # Excel report  (file *or* BytesIO)
    # ------------------------------------------------------------------

    def _build_excel_workbook(self, papers: List[Dict]) -> openpyxl.Workbook:
        """Build an openpyxl Workbook entirely in memory."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PKD Literature"

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 5
        ws.column_dimensions["C"].width = 5
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 25
        ws.column_dimensions["G"].width = 60
        ws.column_dimensions["H"].width = 5
        ws.column_dimensions["I"].width = 5
        ws.column_dimensions["J"].width = 45

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        headers = ["", "", "", "Summary", "Last Author", "Journal", "Key Findings", "", "", "Link"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            if header:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

        for idx, paper in enumerate(papers, start=2):
            ws.cell(row=idx, column=4, value=self.create_summary(paper.get("title", ""), max_words=7))
            ws.cell(row=idx, column=5, value=self.extract_last_author_name(paper.get("authors", "")))
            ws.cell(row=idx, column=6, value=paper.get("journal", ""))
            ws.cell(row=idx, column=7, value=self.create_key_findings(paper.get("title", ""), max_words=20))

            doi = paper.get("doi", "")
            pmid = paper.get("pmid", "")
            link = f"https://doi.org/{doi}" if doi else (f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/" if pmid else "")
            if link:
                link_cell = ws.cell(row=idx, column=10)
                link_cell.value = link
                link_cell.hyperlink = link
                link_cell.font = Font(color="0563C1", underline="single")

        return wb

    def create_excel_report(self, papers: List[Dict], filename: str) -> None:
        """Write Excel report to *filename* on disk."""
        print(f"Creating Excel report: {filename}...")
        wb = self._build_excel_workbook(papers)
        wb.save(filename)
        print(f"  Excel report saved: {filename}")

    def create_excel_bytes(self, papers: List[Dict]) -> bytes:
        """Return the Excel report as raw bytes (for Streamlit downloads)."""
        wb = self._build_excel_workbook(papers)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # PDF report  (file *or* BytesIO)
    # ------------------------------------------------------------------

    def _build_pdf_story(
        self,
        papers: List[Dict],
        categories: Dict,
        pubmed_count: int,
        biorxiv_count: int,
        medrxiv_count: int,
    ) -> list:
        """Return the list of ReportLab flowables for the PDF."""
        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Heading1"], fontSize=18,
            textColor=colors.HexColor("#1f4788"), spaceAfter=30, alignment=TA_CENTER,
        )
        heading1_style = ParagraphStyle(
            "CustomHeading1", parent=styles["Heading1"], fontSize=14,
            textColor=colors.HexColor("#1f4788"), spaceAfter=12, spaceBefore=12,
            borderWidth=1, borderColor=colors.HexColor("#1f4788"), borderPadding=5,
        )
        heading2_style = ParagraphStyle(
            "CustomHeading2", parent=styles["Heading2"], fontSize=12,
            textColor=colors.HexColor("#2E5C8A"), spaceAfter=10, spaceBefore=10,
        )
        body_style = ParagraphStyle(
            "CustomBody", parent=styles["Normal"], fontSize=10, leading=14,
            alignment=TA_JUSTIFY,
        )
        citation_style = ParagraphStyle(
            "Citation", parent=styles["Normal"], fontSize=9, leftIndent=20,
            textColor=colors.HexColor("#444444"), leading=12, spaceAfter=8,
        )

        # Title
        story.append(Paragraph("POLYCYSTIC KIDNEY DISEASE LITERATURE REVIEW", title_style))
        story.append(Paragraph(f"Search Period: {self.start_date} to {self.end_date}", styles["Normal"]))
        story.append(Spacer(1, 20))

        # Summary table
        story.append(Paragraph("SEARCH SUMMARY", heading1_style))
        summary_data = [
            ["PubMed Articles:", str(pubmed_count)],
            ["bioRxiv Preprints:", str(biorxiv_count)],
            ["medRxiv Preprints:", str(medrxiv_count)],
            ["Total Papers:", str(len(papers))],
            ["Date Range:", f"{self.start_date} to {self.end_date}"],
        ]
        t = Table(summary_data, colWidths=[2.5 * inch, 3 * inch])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8EFF7")),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(t)
        story.append(Spacer(1, 20))

        # Notable findings
        story.append(Paragraph("NOTABLE FINDINGS", heading1_style))
        story.append(Spacer(1, 10))

        section_map = [
            ("metabolism", "METABOLISM OR MITOCHONDRIA:", 3),
            ("therapeutics", "THERAPEUTICS:", 4),
            ("cross_species", "CROSS-SPECIES:", 4),
            ("dataset", "NEW DATA SETS:", 3),
        ]
        for key, label, limit in section_map:
            if categories.get(key):
                story.append(Paragraph(label, heading2_style))
                for paper in categories[key][:limit]:
                    story.append(Paragraph(f"&bull; {paper['title']}", body_style))
                    citation = (
                        f'<b>Citation:</b> {paper["authors"]}. '
                        f"{self._paper_link_label(paper)}"
                    )
                    story.append(Paragraph(citation, citation_style))
                story.append(Spacer(1, 15))

        story.append(PageBreak())

        # Complete list
        story.append(Paragraph(f"COMPLETE PAPER LIST ({len(papers)} PAPERS)", heading1_style))
        story.append(Spacer(1, 15))

        for i, paper in enumerate(papers, 1):
            story.append(Paragraph(f'<b>{i}. {paper["title"]}</b>', body_style))
            story.append(
                Paragraph(
                    f'{paper["authors"]}. <i>{paper["journal"]}</i>. {paper["year"]}.',
                    citation_style,
                )
            )
            link_parts = []
            if paper.get("pmid"):
                link_parts.append(
                    f'<link href="https://pubmed.ncbi.nlm.nih.gov/{paper["pmid"]}/" color="blue">'
                    f'PMID: {paper["pmid"]}</link>'
                )
            if paper.get("doi"):
                link_parts.append(
                    f'<link href="https://doi.org/{paper["doi"]}" color="blue">'
                    f'DOI: {paper["doi"]}</link>'
                )
            story.append(Paragraph(" | ".join(link_parts), citation_style))
            story.append(Spacer(1, 10))

        return story

    def create_pdf_report(
        self,
        papers: List[Dict],
        categories: Dict,
        filename: str,
        pubmed_count: int = 0,
        biorxiv_count: int = 0,
        medrxiv_count: int = 0,
    ) -> None:
        """Write PDF report to *filename* on disk."""
        print(f"Creating PDF report: {filename}...")
        doc = SimpleDocTemplate(
            filename, pagesize=letter,
            rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18,
        )
        story = self._build_pdf_story(papers, categories, pubmed_count, biorxiv_count, medrxiv_count)
        doc.build(story)
        print(f"  PDF report saved: {filename}")

    def create_pdf_bytes(
        self,
        papers: List[Dict],
        categories: Dict,
        pubmed_count: int = 0,
        biorxiv_count: int = 0,
        medrxiv_count: int = 0,
    ) -> bytes:
        """Return the PDF report as raw bytes (for Streamlit downloads)."""
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=letter,
            rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18,
        )
        story = self._build_pdf_story(papers, categories, pubmed_count, biorxiv_count, medrxiv_count)
        doc.build(story)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # Deduplication
    # ------------------------------------------------------------------

    @staticmethod
    def _deduplicate_papers(papers: List[Dict]) -> List[Dict]:
        seen_dois: set = set()
        seen_titles: set = set()
        unique: List[Dict] = []
        for paper in papers:
            doi = paper.get("doi", "").strip().lower()
            title = paper.get("title", "").strip().lower()
            if doi and doi in seen_dois:
                continue
            if title and title in seen_titles:
                continue
            if doi:
                seen_dois.add(doi)
            if title:
                seen_titles.add(title)
            unique.append(paper)
        return unique

    # ------------------------------------------------------------------
    # High-level search  (returns structured results)
    # ------------------------------------------------------------------

    def search_all(
        self,
    ) -> Tuple[List[Dict], List[Dict], List[Dict], List[Dict], Dict[str, List[Dict]]]:
        """Run all searches and return (all_papers, pubmed, biorxiv, medrxiv, categories).

        This is the primary entry-point for both the CLI ``run()`` method and
        the Streamlit app.
        """
        self.all_papers = []

        # PubMed
        pmids = self.search_pubmed()
        pubmed_papers = self.get_pubmed_metadata(pmids) if pmids else []
        self.all_papers.extend(pubmed_papers)

        # bioRxiv & medRxiv
        biorxiv_papers = self.search_biorxiv()
        medrxiv_papers = self.search_medrxiv()
        self.all_papers.extend(biorxiv_papers)
        self.all_papers.extend(medrxiv_papers)

        # Deduplicate
        self.all_papers = self._deduplicate_papers(self.all_papers)

        categories = self.categorize_papers(self.all_papers)

        return self.all_papers, pubmed_papers, biorxiv_papers, medrxiv_papers, categories

    # ------------------------------------------------------------------
    # CLI run
    # ------------------------------------------------------------------

    def run(self) -> None:
        """Execute the complete search and write reports to disk (CLI mode)."""
        print("\n" + "=" * 80)
        print("PKD LITERATURE SEARCH")
        print("=" * 80 + "\n")

        all_papers, pubmed_papers, biorxiv_papers, medrxiv_papers, categories = self.search_all()

        print(f"\nTotal papers found: {len(all_papers)}")
        print(f"  PubMed: {len(pubmed_papers)}")
        print(f"  bioRxiv: {len(biorxiv_papers)}")
        print(f"  medRxiv: {len(medrxiv_papers)}")

        if not all_papers:
            print("No papers found for this date range.")
            return

        end_fmt = self.end_date.replace("-", "")
        excel_path = f"{self.output_dir}/{end_fmt}-PKD-Literature-Data.xlsx"
        pdf_path = f"{self.output_dir}/{end_fmt}-PKD-Literature-Summary.pdf"

        self.create_excel_report(all_papers, excel_path)
        self.create_pdf_report(
            all_papers, categories, pdf_path,
            pubmed_count=len(pubmed_papers),
            biorxiv_count=len(biorxiv_papers),
            medrxiv_count=len(medrxiv_papers),
        )

        print("\n" + "=" * 80)
        print("SEARCH COMPLETE")
        print("=" * 80)
        print(f"\nReports generated:")
        print(f"  Excel: {excel_path}")
        print(f"  PDF:   {pdf_path}")
        print(f"\nTotal papers: {len(all_papers)}")
        for cat, label in [
            ("genetics", "Genetics"),
            ("therapeutics", "Therapeutics"),
            ("metabolism", "Metabolism"),
            ("cross_species", "Cross-species"),
            ("dataset", "Datasets"),
        ]:
            print(f"  {label}: {len(categories[cat])}")
        other_count = (
            len(categories["clinical"])
            + len(categories["pathophysiology"])
            + len(categories["other"])
        )
        print(f"  Other: {other_count}")


# ======================================================================
# CLI entry point
# ======================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Search PKD literature and generate reports",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python pkd_literature_search.py
  python pkd_literature_search.py --start 2026-02-01 --end 2026-02-08
  python pkd_literature_search.py --start 2026-02-01 --end 2026-02-08 --output /tmp
        """,
    )
    parser.add_argument("--start", type=str, help="Start date (YYYY-MM-DD). Default: 7 days ago")
    parser.add_argument("--end", type=str, help="End date (YYYY-MM-DD). Default: today")
    parser.add_argument("--output", type=str, default=".", help="Output directory. Default: current directory")
    args = parser.parse_args()

    if not args.end:
        args.end = datetime.now().strftime("%Y-%m-%d")
    if not args.start:
        end_date = datetime.strptime(args.end, "%Y-%m-%d")
        args.start = (end_date - timedelta(days=7)).strftime("%Y-%m-%d")

    try:
        datetime.strptime(args.start, "%Y-%m-%d")
        datetime.strptime(args.end, "%Y-%m-%d")
    except ValueError:
        print("Error: Dates must be in YYYY-MM-DD format")
        return 1

    searcher = PKDLiteratureSearch(args.start, args.end, args.output)
    searcher.run()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
